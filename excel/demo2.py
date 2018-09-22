#2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写
import openpyxl

if __name__ == "__main__":
    #write file
    # wb = openpyxl.Workbook()
    # sheet = wb.active
    # sheet.title = "2007测试表"
    # value = [["名称", "价格", "出版社", "语言"],
    #          ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
    #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
    #          ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    # for i in range(len(value)):
    #     for j in range(len(value[i])):
    #         sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    #
    # wb.save("2007_test.xlsx")

    #read file
    wb = openpyxl.load_workbook("2007_test.xlsx")
    print(wb.sheetnames)
    sheet = wb.active
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()